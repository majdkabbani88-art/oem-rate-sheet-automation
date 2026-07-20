"""
OEM Interest Rate Sheet Automation
------------------------------------
Every finance cycle, each OEM (vehicle manufacturer's captive finance arm)
sends over their current consumer lending rates by term (24/36/48/60/72
months). These sheets need to end up in the internal system as a clean,
standardized table before the data entry team can act on them.

The problem this tool solves: every OEM formats their sheet differently.
Some flag what changed, some don't. Some deliver Excel, some deliver PDF.
Some list rates in a simple table, some in a matrix, some bury the table
under header notes with merged cells. None of that is consistent, and it
used to mean manually opening each file and eyeballing it against last
cycle's numbers.

This tool:
    1. Reads every OEM's rate sheet from INPUT_DIR, using a dedicated
       parser per OEM format (because no two formats are alike).
    2. Normalizes everything to one common structure: OEM, Model,
       Term (Months), Rate (%).
    3. Compares the result against last cycle's saved snapshot
       (REFERENCE_DIR) to detect what changed, what's new, and what
       was dropped.
    4. Writes one clean Excel workbook to OUTPUT_DIR with four sheets:
         - Clean_Data       : the full, standardized rate table, ready
                               for the data entry team
         - Changes          : rows where the rate moved since last
                               cycle, old vs. new value, highlighted red
         - Added_Vehicles   : models that are brand new this cycle
         - Removed_Vehicles : models that disappeared from an OEM's
                               sheet since last cycle
    5. Saves the new Clean_Data as this cycle's reference snapshot,
       so next cycle's comparison has something to diff against.

Adding a new OEM later just means writing one new parser function and
registering it in OEM_PARSERS below — nothing else in the pipeline changes.

Author: Majd Kabbani
"""

import os
from datetime import date

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_DIR = "input"
REFERENCE_DIR = "reference"
OUTPUT_DIR = "output"

REFERENCE_FILE = os.path.join(REFERENCE_DIR, "previous_clean_data.xlsx")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"clean_rates_{date.today().isoformat()}.xlsx")

COLUMNS = ["OEM", "Model", "Term_Months", "Rate_Percent"]


# ---------------------------------------------------------------------------
# OEM-SPECIFIC PARSERS
# Each OEM formats their sheet differently, so each gets its own parser.
# Every parser's only job is to return a clean DataFrame with columns:
# OEM, Model, Term_Months, Rate_Percent
# ---------------------------------------------------------------------------

def parse_meridian(path: str) -> pd.DataFrame:
    """Simple table format. Has its own 'Change Flag' column, but we still
    compute changes independently against the reference so every OEM is
    treated consistently regardless of whether they self-report changes."""
    df = pd.read_excel(path, header=3)  # first 3 rows are title/date/blank
    df = df.rename(columns={"Term (Months)": "Term_Months", "Rate (%)": "Rate_Percent"})
    df["OEM"] = "Meridian Motors"
    return df[COLUMNS]


def parse_atlas(path: str) -> pd.DataFrame:
    """Matrix format: one row per model, one column per term. No change
    flag at all — needs to be melted into long format."""
    df = pd.read_excel(path)
    term_cols = [c for c in df.columns if c != "Model"]
    long_df = df.melt(id_vars="Model", value_vars=term_cols,
                       var_name="Term_Months", value_name="Rate_Percent")
    long_df["Term_Months"] = long_df["Term_Months"].str.replace(" mo", "", regex=False).astype(int)
    long_df["OEM"] = "Atlas Auto Finance"
    return long_df[COLUMNS]


def parse_northbridge(path: str) -> pd.DataFrame:
    """Messy layout: 3 header/notice rows before the real table starts,
    and the model name is only present on the first term row per model
    (blank on the rest, since it's visually 'grouped' in the source file).
    Needs a forward-fill to reconstruct the model for every row."""
    df = pd.read_excel(path, header=None, skiprows=4)
    df.columns = ["Model", "Term_Months", "Rate_Percent"]
    df["Model"] = df["Model"].ffill()
    df["OEM"] = "Northbridge Capital"
    return df[COLUMNS]


def parse_crestline_pdf(path: str) -> pd.DataFrame:
    """PDF source. Table is extracted page by page with pdfplumber rather
    than read directly like the Excel-based OEMs."""
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            headers, *data_rows = table
            for row in data_rows:
                rows.append(dict(zip(headers, row)))
    df = pd.DataFrame(rows)
    df = df.rename(columns={"Term (Months)": "Term_Months", "Rate (%)": "Rate_Percent"})
    df["Term_Months"] = df["Term_Months"].astype(int)
    df["Rate_Percent"] = df["Rate_Percent"].astype(float)
    df["OEM"] = "Crestline Financial"
    return df[COLUMNS]


# Registry: filename pattern -> parser function.
# Adding a new OEM later means adding one line here.
OEM_PARSERS = {
    "meridian_current.xlsx": parse_meridian,
    "atlas_current.xlsx": parse_atlas,
    "northbridge_current.xlsx": parse_northbridge,
    "crestline_current.pdf": parse_crestline_pdf,
}


# ---------------------------------------------------------------------------
# PIPELINE
# ---------------------------------------------------------------------------
def load_all_current_sheets() -> pd.DataFrame:
    frames = []
    for filename, parser in OEM_PARSERS.items():
        path = os.path.join(INPUT_DIR, filename)
        if not os.path.exists(path):
            print(f"  [skip] {filename} not found in input folder")
            continue
        df = parser(path)
        frames.append(df)
        print(f"  [ok] {filename}: {len(df)} rate rows parsed")
    return pd.concat(frames, ignore_index=True)


def compare_to_reference(current: pd.DataFrame, previous: pd.DataFrame) -> dict:
    key_cols = ["OEM", "Model", "Term_Months"]
    merged = current.merge(previous, on=key_cols, how="outer",
                            suffixes=("_current", "_previous"), indicator=True)

    # Changed: present in both, but rate differs
    both = merged[merged["_merge"] == "both"].copy()
    changed = both[both["Rate_Percent_current"] != both["Rate_Percent_previous"]].copy()
    changed = changed.rename(columns={
        "Rate_Percent_previous": "Previous Rate (%)",
        "Rate_Percent_current": "New Rate (%)",
    })[["OEM", "Model", "Term_Months", "Previous Rate (%)", "New Rate (%)"]]

    # Added vehicles: models present in current for an OEM, absent in previous for that OEM
    current_models = set(current[["OEM", "Model"]].itertuples(index=False, name=None))
    previous_models = set(previous[["OEM", "Model"]].itertuples(index=False, name=None))
    added_models = current_models - previous_models
    removed_models = previous_models - current_models

    added_df = pd.DataFrame(sorted(added_models), columns=["OEM", "Model"])
    removed_df = pd.DataFrame(sorted(removed_models), columns=["OEM", "Model"])

    return {"changed": changed, "added": added_df, "removed": removed_df}


# ---------------------------------------------------------------------------
# OUTPUT WORKBOOK
# ---------------------------------------------------------------------------
def style_header(ws, ncols, row=1):
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, ncols, width=18):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def write_df(ws, df: pd.DataFrame, highlight_col=None):
    headers = list(df.columns)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, len(headers))
    red_fill = PatternFill("solid", fgColor="FDE7E7")
    red_font = Font(color="B5493A", bold=True)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row[h])
            if highlight_col and h == highlight_col:
                cell.fill = red_fill
                cell.font = red_font
    autosize(ws, len(headers))


def build_output_workbook(clean_df, diffs):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Clean_Data"
    write_df(ws, clean_df.sort_values(["OEM", "Model", "Term_Months"]))

    ws = wb.create_sheet("Changes")
    if not diffs["changed"].empty:
        write_df(ws, diffs["changed"], highlight_col="New Rate (%)")
    else:
        ws["A1"] = "No rate changes this cycle."

    ws = wb.create_sheet("Added_Vehicles")
    if not diffs["added"].empty:
        write_df(ws, diffs["added"])
    else:
        ws["A1"] = "No new vehicles this cycle."

    ws = wb.create_sheet("Removed_Vehicles")
    if not diffs["removed"].empty:
        write_df(ws, diffs["removed"])
    else:
        ws["A1"] = "No vehicles removed this cycle."

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("Parsing current-cycle OEM rate sheets...")
    current = load_all_current_sheets()

    previous = pd.read_excel(REFERENCE_FILE)

    print("Comparing against last cycle's reference snapshot...")
    diffs = compare_to_reference(current, previous)

    print("Writing clean output workbook...")
    build_output_workbook(current, diffs)

    # This cycle's clean data becomes next cycle's reference
    current.to_excel(REFERENCE_FILE, index=False)

    print(f"\nDone. Report saved to: {OUTPUT_FILE}")
    print(f"  Total rate rows: {len(current)}")
    print(f"  Changed: {len(diffs['changed'])}")
    print(f"  Added vehicles: {len(diffs['added'])}")
    print(f"  Removed vehicles: {len(diffs['removed'])}")


if __name__ == "__main__":
    main()
